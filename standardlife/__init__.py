import openpyxl
from utils import clean_spreadsheet, fetch_with_backoff, get_random_user_agent, get_xlsx_filepath


def standard_life_runner() -> None:
    out_xlsx = get_xlsx_filepath("standard_life.xlsx")
    clean_spreadsheet(out_xlsx)

    headers = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "en-US,en;q=0.9,fr-FR;q=0.8,fr;q=0.7",
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "Cookie": 'JSESSIONID="30EB783BFD98F70314B78BC2EDC5C448.P1FD07H02:P1FD07H02_FUND_FILTER"; TS01c6f0a9=01f942b434b3d5590a698b71e2202d2d83f11e65ee748a116296c20aa75463e76a75fd93f857388bdb784ba4ba6f66389c114df3b0dfc6517a5f940768f626da7cdbf1f09f; visid_incap_2552455=wxAwufPRTFKW68Hv7cfo37pV0WEAAAAAQUIPAAAAAACCCAGdWHaD9S1mtRfwZ9Yu; dtCookie=4$888676291CF41C65B3376CD4C1967BE3|187ac967c3e45871|1; TS01dedb5e=01f942b434a15e5a0926a824cf9f01cba616572424748a116296c20aa75463e76a75fd93f8b8b9ff7a76523f53fc9e9004721841e584039ea856ce17e6f5483ae9f64a9305; sl#cookiepreferences={"cookiecategoryids":"1|2|3","version":"2"}; incap_ses_1100_2552455=FVmzNCHo82+RTOBqrvxDD9IH0mEAAAAAyVy8rMnNeejvBsZ+Z6//tg==; incap_ses_1105_2552455=iGDlLDYkxxHZBsCnxsZVD9MH0mEAAAAAaJ7hPVsTQNyNQZjEtpKc1w==; JSESSIONID=9FD86719F8D492EBC40D12DD5B1B69E5; TS019def75=01f942b434072c47fbcea1557cb63859ca8ba3fac8d39924d28066f499f41057ac6f14a7f75c7138e0cabe1b26f04da43a45d95fc7427d94e11b5afa8046ce0739753d3162; TS0122a24a=01f942b4342dc173601d2f5012d1bc5871626c064f2458c21b40babbe93d99feba3b865401b745e0334b28f190ed7d13db32076b01',
        "Host": "secure.standardlife.co.uk",
        "Pragma": "no-cache",
        "Upgrade-Insecure-Requests": "1",
    }
    headers.update(get_random_user_agent())

    payload = {
        "_": 1641154486267
    }

    url = 'https://secure.standardlife.co.uk/secure/fundfilter/rest/results/funds/INVESTMENTS/PPORT/FIRST_NZ/existingcustomer?_=1641154486267'
    res = fetch_with_backoff(url, headers=headers, data=payload)
    # res = requests.get(url, headers=headers, data=payload)
    print('[#] Scraping Standard Life.')
    if res:
        obj = res.json()
        list_funds = obj["aaData"]
        sorted_funds = sorted(list_funds, key=lambda x: x[0])
        xlsx_iter = 2
        wb = openpyxl.load_workbook(out_xlsx)
        ws = wb['standard']
        print(f'[#] found {len(sorted_funds)} funds. [#]')
        for fund in sorted_funds:
            fund_url = f"https://digital.feprecisionplus.com/documents/standardlifepat/en-GB/{fund[17]}/KI"
            ws.cell(xlsx_iter, 1).value = fund[0]
            ws.cell(xlsx_iter, 2).value = fund[18]
            c = ws.cell(xlsx_iter, 3, fund_url)
            c.hyperlink = fund_url
            c.style = "Hyperlink"
            if xlsx_iter % 50 == 0:
                wb.save(out_xlsx)
            xlsx_iter += 1
        wb.save(out_xlsx)
        wb.close()
